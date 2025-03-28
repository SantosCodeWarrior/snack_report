from django.shortcuts import render
from django.shortcuts import render,render_to_response
from snacks import models
from django.http import HttpResponse
import json
from django.db.models import Count
from datetime import date,timedelta,datetime
from collections import Counter
import numpy as np
import decimal
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.http import HttpResponseRedirect
from django.core.urlresolvers import reverse
import os
import time
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
from num2words import num2words

#####################################################
import smtplib
from django.core.mail.message import EmailMessage
from django.core.mail import send_mail
from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEText import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import os,sys
from os import path
import re
import sys, ast
import subprocess
from subprocess import Popen, PIPE
import pprint
from django.core.mail import EmailMultiAlternatives
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
import requests
import re
from bs4 import BeautifulSoup
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import pdfkit
import webbrowser
from django.db.models import Max
from django.db.models import Min
from shutil import copyfile
import shutil
from django.db.models import Sum
from time import strptime
from calendar import monthrange
from django.db import connection	

##############################################
import xlsxwriter
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
try: 
	from openpyxl.cell import get_column_letter
except ImportError:
	from openpyxl.utils import get_column_letter
import os
import time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from shutil import copyfile
import shutil
from openpyxl.styles import Style
from django.views.decorators.csrf import csrf_exempt
from num2words import num2words
#####################################################


@csrf_exempt
def export_tracker(request):
	date_filter   = json.loads(request.POST['date_filter'])
	select_dir    = os.path.dirname(__file__)
	srcfile       = select_dir+'/static/template/report.xlsx'
	dstroot       = select_dir+'/Export/report.xlsx'	 
	copyfile(srcfile, dstroot)
	my_dir 	      = os.path.dirname(__file__)		
	location      = (my_dir+'/Export/report.xlsx')
	wb 	   	      = load_workbook(my_dir+'/Export/report.xlsx')
	ws     	      = wb.get_sheet_by_name("Sheet1")	
	thin_border   = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	my_style      = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	sums = 0
	ii   = 8
	j 	 = 1

	snacks_det = models.owner_master.objects.filter(status=0).first()


	snacks_details = models.journal_master.objects.filter(entry_date=date_filter)
	ws.cell('A1').value = snacks_det.shop_name
	ws.cell('C2').value = snacks_det.owner_name
	ws.cell('C3').value = snacks_det.contact_no
	ws.cell('D2').value = snacks_det.address

	for t in snacks_details:	
	 	thin_border    = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	 	my_style       = Style(font=Font(name='Calibri', size=8, bold=False),border=thin_border)
	 	_row_style     = Style(font=Font(name='Calibri', size=8, bold=True),border=thin_border)
		ws.cell('A'+str(ii)).value = j
		ws.cell('B'+str(ii)).value = t.entry_date
		ws.cell('C'+str(ii)).value = t.staff.staff_master
		ws.cell('D'+str(ii)).value = t.standard.items
		ws.cell('E'+str(ii)).value = t.standard.rate
		ws.cell('F'+str(ii)).value = t.qty
		ws.cell('G'+str(ii)).value = t.amount
		sums+=t.amount
		ws.cell('A'+str(ii)).style = _row_style	 
		ws.cell('B'+str(ii)).style = _row_style	 
		ws.cell('C'+str(ii)).style = _row_style	 
		ws.cell('D'+str(ii)).style = _row_style	 
		ws.cell('E'+str(ii)).style = _row_style	 
		ws.cell('F'+str(ii)).style = _row_style
		ws.cell('G'+str(ii)).style = _row_style
		ii+=1
		j+=1
	_row_count 		= ii
	_border	    = Style(font=Font(name='Arial', size=9, italic=False,bold=True),alignment=Alignment(horizontal='right'))
	_font_style	= Style(font=Font(name='Arial', size=9, italic=False,bold=False),alignment=Alignment(horizontal='right'))
	merge 		= ('E'+str(_row_count+2)+':G'+str(_row_count+2))
	merge1 		= ('E'+str(_row_count+3)+':G'+str(_row_count+3))
	merge2 		= ('E'+str(_row_count+4)+':G'+str(_row_count+4))
	print '---',merge,'----',merge1,'----',merge2
	ws.merge_cells(merge)
	ws.merge_cells(merge1)
	ws.merge_cells(merge2)
	ws.cell('E'+str(_row_count+2)).value = "For "+str(snacks_det.shop_name)
	ws.cell('G'+str(_row_count+1)).value = sums	
	ws.cell('E'+str(_row_count+3)).value = "Proprietor :"+str(snacks_det.owner_name)
	ws.cell('E'+str(_row_count+4)).value = "Sign:-"
	ws.cell('E'+str(_row_count+2)).style = _border
	ws.cell('G'+str(_row_count+2)).style = _border
	ws.cell('E'+str(_row_count+3)).style = _border
	ws.cell('E'+str(_row_count+4)).style = _border	
	ws.cell('G'+str(_row_count+1)).style = _border	

	wb.save(my_dir+'/Export/report.xlsx')
	return HttpResponse(json.dumps('done'))

def report_export(request):
	import os, tempfile, zipfile
	from django.core.servers.basehttp import FileWrapper
	from django.conf import settings
	import mimetypes
	my_dir 	 	   					= os.path.dirname(__file__)
	location 	   					= (my_dir+'/Export/report.xlsx')
	filename 	   					= my_dir+'/Export/report.xlsx' #Select your file here.
	download_name  					= "Report.xlsx"
	wrapper        					= FileWrapper(open(filename))
	content_type   					= mimetypes.guess_type(filename)[0]
	response       					= HttpResponse(wrapper,content_type=content_type)
	response['Content-Length']      = os.path.getsize(filename)
	response['Content-Disposition'] = "attachment; filename=%s"%download_name
	return response